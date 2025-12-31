from openpyxl import load_workbook
import requests
import json 

CITY_FILE_NAME="backend/weather/AMap_adcode_citycode.xlsx"
SHEET_NAME="Sheet1"

def GetCityAdcode(city) -> int:
    wb = load_workbook(CITY_FILE_NAME)
    ws = wb[SHEET_NAME] 
    for row in ws.iter_rows(values_only=True):
        if row[0] == str(city):
            return row[1]

def GetWeatherApiKey() -> str:
    return open("backend/weather/.api_key").read()

def WeatherInformation(city) -> any:
    url = "https://restapi.amap.com/v3/weather/weatherInfo?city={city}&key={key}".format(city=GetCityAdcode(city), key=GetWeatherApiKey())
    ret = requests.get(url)
    return ret.json()
